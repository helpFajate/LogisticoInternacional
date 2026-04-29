import pyodbc
import pandas as pd
from fpdf import FPDF
from fpdf.enums import XPos, YPos
import warnings
import os
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

warnings.filterwarnings('ignore', category=UserWarning)

# ---------------------------------------------------------
# CLASE PDF
# ---------------------------------------------------------
class ReporteEmpaque(FPDF):
    def __init__(self, datos_encabezado, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self.datos_envio = datos_encabezado
        self.ruta_base_img = r"C:\Users\mesa_ayuda\Desktop\Internacional\imagenes"

    def header(self):
        if self.page_no() == 1:
            y_inicial = 10
            alto_celda_cuadro = 6
            alto_total_cuadro = alto_celda_cuadro * 3
            alto_imagenes = alto_total_cuadro * 0.95

            x_cuadro = 130
            ancho_cuadro = 70
            self.set_y(y_inicial)
            self.set_x(x_cuadro)
            self.set_font("helvetica", "", 10)
            self.set_text_color(0, 0, 0)
            self.cell(ancho_cuadro, alto_celda_cuadro, "PACKING LIST/LISTA DE EMPAQUE", border=1, align="C",
                      new_x=XPos.LMARGIN, new_y=YPos.NEXT)
            self.set_x(x_cuadro)
            factura = self.datos_envio.get('factura', 'N/A')
            self.cell(ancho_cuadro, alto_celda_cuadro, f"Fact. N°: {factura}", border=1, align="C",
                      new_x=XPos.LMARGIN, new_y=YPos.NEXT)
            self.set_x(x_cuadro)
            self.cell(ancho_cuadro, alto_celda_cuadro, f"Pagina N°: {self.page_no()}", border=1, align="C",
                      new_x=XPos.LMARGIN, new_y=YPos.NEXT)

            y_img = y_inicial + 0.45
            img_vivell   = os.path.join(self.ruta_base_img, "vivell.png")
            img_invima   = os.path.join(self.ruta_base_img, "INVIMA.png")
            img_operador = os.path.join(self.ruta_base_img, "OPERADOR.png")
            if os.path.exists(img_vivell):   self.image(img_vivell,   x=10, y=y_img,     h=alto_imagenes)
            if os.path.exists(img_invima):   self.image(img_invima,   x=60, y=y_img + 3, w=30)
            if os.path.exists(img_operador): self.image(img_operador, x=95, y=y_img + 3, w=25)

            self.ln(12)
            self.set_font("helvetica", "", 8)
            w_cols = [35, 60, 35, 60]
            alto_fila = 6

            self.cell(w_cols[0] + w_cols[1], 5, "Customer/Cliente", border=1, align="C")
            self.cell(w_cols[2] + w_cols[3], 5, "Vendor/Vendedor",  border=1, align="C",
                      new_x=XPos.LMARGIN, new_y=YPos.NEXT)

            def fila_datos(label1, val1, label2, val2):
                curr_x, curr_y = self.get_x(), self.get_y()
                self.cell(w_cols[0], alto_fila, "", border=1)
                self.cell(w_cols[1], alto_fila, "", border=1)
                self.cell(w_cols[2], alto_fila, "", border=1)
                self.cell(w_cols[3], alto_fila, "", border=1, new_x=XPos.LMARGIN, new_y=YPos.NEXT)
                self.set_xy(curr_x, curr_y + 0.5)
                self.multi_cell(w_cols[0], 3, label1, align="L")
                self.set_xy(curr_x + w_cols[0], curr_y + 0.5)
                self.multi_cell(w_cols[1], 3, str(val1), align="L")
                self.set_xy(curr_x + w_cols[0] + w_cols[1], curr_y + 0.5)
                self.multi_cell(w_cols[2], 3, label2, align="L")
                self.set_xy(curr_x + w_cols[0] + w_cols[1] + w_cols[2], curr_y + 0.5)
                self.multi_cell(w_cols[3], 3, str(val2), align="L")
                self.set_xy(curr_x, curr_y + alto_fila)

            d = self.datos_envio
            fila_datos("Name/Nombre",       d.get('cliente_nombre',''), "Name/Nombre",       d.get('cia_nombre',''))
            fila_datos("ID/Nit",            d.get('cliente_nit',''),    "ID/NIT",            d.get('cia_nit',''))
            fila_datos("Address/Dirección", d.get('cliente_dir',''),    "Address/Dirección", d.get('cia_dir',''))
            fila_datos("City/Ciudad",       d.get('cliente_ciu',''),    "City/Ciudad",       d.get('cia_ciu',''))
            fila_datos("State/Depto",       d.get('cliente_dep',''),    "State/Depto",       d.get('cia_dep',''))
            fila_datos("Country/País",      d.get('cliente_pais',''),   "Country/País",      d.get('cia_pais',''))
            fila_datos("Phone/Teléfono",    d.get('cliente_tel',''),    "Phone/Teléfono",    d.get('cia_tel',''))
            fila_datos("Contact/Contacto",  d.get('cliente_cont',''),   "Vendor/Vendedor",   d.get('vendedor',''))
            self.ln(5)
        else:
            self.set_y(10)

    def footer(self):
        self.set_y(-15)
        self.set_font("helvetica", "I", 8)
        self.set_text_color(0, 0, 0)
        self.cell(0, 10, f"Página {self.page_no()}/{{nb}}", align="C")

    def check_page_break(self, h):
        if self.get_y() + h > self.page_break_trigger:
            self.add_page()


# ---------------------------------------------------------
# HELPERS DE ESTILO EXCEL
# ---------------------------------------------------------
def _border_all():
    s = Side(style='thin', color='000000')
    return Border(left=s, right=s, top=s, bottom=s)

def _border_top():
    s = Side(style='thin', color='000000')
    return Border(top=s)

def _fill(hex_color):
    return PatternFill('solid', start_color=hex_color, end_color=hex_color)

def _font(bold=False, size=9, color='000000', italic=False):
    return Font(name='Arial', bold=bold, size=size, color=color, italic=italic)

def _align(h='left', v='center', wrap=True):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)


# ---------------------------------------------------------
# CONEXIÓN / DATOS
# ---------------------------------------------------------
def _obtener_datos(cia, tipo_docto, consec_inicial, consec_final):
    conn_str = (
        r'DRIVER={ODBC Driver 17 for SQL Server};'
        r'SERVER=AMATERASU\SIESA;'
        r'DATABASE=Reportes;'
        r'Trusted_Connection=yes;'
    )
    try:
        conn = pyodbc.connect(conn_str)
        query = (
            "EXEC sp_ti_listaEmpaque "
            "@v_cia=?, @v_idco='063', @v_idtipodocto=?, "
            "@v_consecinicial=?, @v_consecfinal=?"
        )
        df = pd.read_sql(query, conn, params=(cia, tipo_docto, consec_inicial, consec_final))
        conn.close()
        return df
    except Exception as e:
        print(f"Error conexión: {e}")
        return pd.DataFrame()


def _datos_encabezado(f):
    return {
        'factura':         f"{f['f_prefijo']}-{f['f_consec_docto']}",
        'cliente_nombre':  f['f_cliente_razon_soc'],
        'cliente_nit':     f['f_cliente_nit'],
        'cliente_dir':     f['f_direccion1_cliente'],
        'cliente_ciu':     f['f_ciudad_cliente'],
        'cliente_dep':     f['f_depto_cliente'],
        'cliente_pais':    f['f_pais_cliente'],
        'cliente_tel':     f['f_telefono_cliente'],
        'cliente_cont':    f['f_contacto_cliente'],
        'cia_nombre':      f['f_razon_social_cia'],
        'cia_nit':         f['f_nit_cia'],
        'cia_dir':         f['f_direccion1_co'],
        'cia_ciu':         f['f_ciudad_co'],
        'cia_dep':         f['f_depto_co'],
        'cia_pais':        f['f_pais_co'],
        'cia_tel':         f['f_telefono_co'],
        'vendedor':        f['f_vendedor_razon_social'],
    }


# ---------------------------------------------------------
# GENERADOR PDF
# ---------------------------------------------------------
def generar_pdf_empaque(cia, tipo_docto, consec_inicial, consec_final):
    df = _obtener_datos(cia, tipo_docto, consec_inicial, consec_final)
    if df.empty:
        return None, None

    datos_h = _datos_encabezado(df.iloc[0])
    w = [18, 12, 12, 64, 64, 20]

    pdf = ReporteEmpaque(datos_encabezado=datos_h)
    pdf.alias_nb_pages()
    pdf.add_page()

    total_gral_cant = 0
    total_gral_pb   = 0.0

    for caja, items in df.groupby('Caja', sort=True):
        val_pb = float(items['f2_ent_peso_bruto'].iloc[0]) if pd.notna(items['f2_ent_peso_bruto'].iloc[0]) else 0.0
        val_pn = float(items['f2_ent_peso_neto'].iloc[0])  if pd.notna(items['f2_ent_peso_neto'].iloc[0])  else 0.0

        pdf.check_page_break(20)
        pdf.set_font("helvetica", "B", 9)
        pdf.set_fill_color(240, 240, 240)
        pdf.set_text_color(0, 0, 0)
        pdf.cell(sum(w), 8,
                 f" Box/Caja N°: {caja}  - Remision/Remission: {items['f2_remision'].iloc[0]}",
                 border=1, fill=True, new_x=XPos.LMARGIN, new_y=YPos.NEXT)

        # Encabezado columnas
        pdf.set_font("helvetica", "B", 8)
        pdf.set_fill_color(220, 220, 220)
        titulos    = ["Reference", "Size/Talla", "Colour/Color", "Description", "Composition", "Qty/Cant"]
        h_enc      = 8
        curr_x, curr_y = pdf.get_x(), pdf.get_y()
        for i, t in enumerate(titulos):
            pdf.set_xy(curr_x, curr_y)
            pdf.cell(w[i], h_enc, "", border=1, fill=True)
            texto = t.replace("/", "/\n") if w[i] <= 20 else t
            y_m = curr_y + 1 if '\n' in texto else curr_y + 2.5
            pdf.set_xy(curr_x, y_m)
            pdf.multi_cell(w[i], 3, texto, align="C")
            curr_x += w[i]
        pdf.set_xy(10, curr_y + h_enc)

        # Filas
        pdf.set_font("helvetica", "", 7)
        cant_caja = 0.0
        for _, row in items.iterrows():
            txt_desc = str(row['f2_desc_crit1'])
            txt_comp = str(row['f2_desc_crit2'])
            l_desc   = (pdf.get_string_width(txt_desc) // w[3]) + 1
            l_comp   = (pdf.get_string_width(txt_comp) // w[4]) + 1
            h_fila   = max(7, max(l_desc, l_comp) * 4)

            pdf.check_page_break(h_fila)
            curr_y = pdf.get_y()
            pdf.cell(w[0], h_fila, str(row['f2_referencia']),  border=1, align="C")
            pdf.cell(w[1], h_fila, str(row['f2_ext2_det']),    border=1, align="C")
            pdf.cell(w[2], h_fila, str(row['f2_id_ext1_det']), border=1, align="C")
            x_desc = pdf.get_x()
            pdf.multi_cell(w[3], h_fila / l_desc if l_desc > 0 else h_fila, txt_desc, border=1)
            pdf.set_xy(x_desc + w[3], curr_y)
            pdf.multi_cell(w[4], h_fila / l_comp if l_comp > 0 else h_fila, txt_comp, border=1)
            pdf.set_xy(x_desc + w[3] + w[4], curr_y)
            v_can = float(row['f2_cantidad'])
            pdf.cell(w[5], h_fila, f"{v_can:.0f}", border=1, align="C",
                     new_x=XPos.LMARGIN, new_y=YPos.NEXT)
            cant_caja += v_can

        # Subtotal caja
        pdf.set_font("helvetica", "B", 8)
        pdf.cell(w[0]+w[1]+w[2], 7, f"Total BOX/CAJA: {caja}",                    border="LTB")
        pdf.cell(w[3],           7, f"Gross Weight/Peso Bruto: {val_pb:.2f}",      border="TB")
        pdf.cell(w[4],           7, f"Net Weight/Peso Neto: {val_pn:.2f}",         border="TB", align="R")
        pdf.cell(w[5],           7, f"{cant_caja:.0f}",                            border="RTB", align="C",
                 new_x=XPos.LMARGIN, new_y=YPos.NEXT)
        pdf.ln(4)

        total_gral_cant += cant_caja
        total_gral_pb   += val_pb

    # Gran total
    pdf.check_page_break(15)
    pdf.set_font("helvetica", "B", 10)
    pdf.set_fill_color(200, 200, 200)
    pdf.set_text_color(0, 0, 0)
    pdf.cell(sum(w[:4]), 10, "  TOTAL GENERAL", border=1, fill=True)
    pdf.cell(sum(w[4:]), 10,
             f"{total_gral_cant:.0f} UNIDADES / {total_gral_pb:.2f} KG",
             border=1, fill=True, align="C")

    try:
        pdf_bytes     = pdf.output()
        nombre_archivo = f"Packing_List_{consec_inicial}.pdf"
        return pdf_bytes, nombre_archivo
    except Exception as e:
        print(f"Error al generar PDF: {e}")
        return None, None


# ---------------------------------------------------------
# GENERADOR EXCEL  (idéntico al Excel de referencia LE_IN63)
# ---------------------------------------------------------
def generar_excel_empaque(cia, tipo_docto, consec_inicial, consec_final):
    from openpyxl.drawing.image import Image as XLImage
    from openpyxl.styles import PatternFill as PF

    df = _obtener_datos(cia, tipo_docto, consec_inicial, consec_final)
    if df.empty:
        return None, None

    d        = _datos_encabezado(df.iloc[0])
    wb       = Workbook()
    ws       = wb.active
    ws.title = "Packing List"

    # Anchos exactos del Excel de referencia
    ws.column_dimensions['A'].width = 23.57
    ws.column_dimensions['B'].width = 9.57
    ws.column_dimensions['C'].width = 13.0
    ws.column_dimensions['D'].width = 34.29
    ws.column_dimensions['E'].width = 24.86
    ws.column_dimensions['F'].width = 34.57

    # Helpers de estilo
    def b_all():
        s = Side(style='thin', color='000000')
        return Border(left=s, right=s, top=s, bottom=s)
    def b_left_top():
        s = Side(style='thin', color='000000')
        return Border(left=s, top=s)
    def b_ltr():   # left+top+right
        s = Side(style='thin', color='000000')
        return Border(left=s, top=s, right=s)
    def b_lr():    # left+right only
        s = Side(style='thin', color='000000')
        return Border(left=s, right=s)
    def b_lbr():   # left+bottom+right
        s = Side(style='thin', color='000000')
        return Border(left=s, bottom=s, right=s)
    def b_left():
        s = Side(style='thin', color='000000')
        return Border(left=s)
    def b_right():
        s = Side(style='thin', color='000000')
        return Border(right=s)
    def b_bottom():
        s = Side(style='thin', color='000000')
        return Border(bottom=s)
    def b_lb():
        s = Side(style='thin', color='000000')
        return Border(left=s, bottom=s)
    def b_rb():
        s = Side(style='thin', color='000000')
        return Border(right=s, bottom=s)

    fnt  = lambda bold=False, size=11: Font(name='Calibri', bold=bold, size=size)
    aln  = lambda h='left', wrap=False: Alignment(horizontal=h, wrap_text=wrap)

    # ── ZONA LOGOS (filas 1-6) ─────────────────────────
    for r in range(1, 7):
        ws.row_dimensions[r].height = 15

    # F2:F3 merged — PACKING LIST
    ws.merge_cells('F2:F3')
    ws['F2'].value     = 'PACKING LIST/LISTA DE EMPAQUE'
    ws['F2'].font      = Font(name='Calibri', size=12)
    ws['F2'].alignment = Alignment(horizontal='center', wrap_text=True)
    ws['F2'].border    = b_all()

    # F4 — Factura
    ws['F4'].value     = f"Fact. N°: {d['factura']}"
    ws['F4'].font      = fnt(size=11)
    ws['F4'].alignment = aln('center')
    ws['F4'].border    = b_all()

    # F5 — Página
    ws['F5'].value     = 'Pagina N°: 1'
    ws['F5'].font      = fnt(size=11)
    ws['F5'].alignment = aln('center')
    ws['F5'].border    = b_all()

    # Logos
    ruta_img = r"C:\Users\mesa_ayuda\Desktop\Internacional\imagenes"
    for nombre_img, celda, ancho_px, alto_px in [
        ("vivell.png",    'A1', 160, 55),
        ("INVIMA.png",    'C1',  95, 45),
        ("OPERADOR.png",  'D1',  85, 45),
    ]:
        ruta = os.path.join(ruta_img, nombre_img)
        if os.path.exists(ruta):
            try:
                img = XLImage(ruta)
                img.width  = ancho_px
                img.height = alto_px
                ws.add_image(img, celda)
            except Exception:
                pass

    # ── FILA 7: Costumer/Cliente | Vendor/Vendedor ──────
    ws.merge_cells('A7:D7')
    ws['A7'].value     = 'Costumer/Cliente'
    ws['A7'].font      = fnt(size=11)
    ws['A7'].alignment = aln('center')
    ws['A7'].border    = b_all()

    ws.merge_cells('E7:F7')
    ws['E7'].value     = 'Vendor/Vendedor'
    ws['E7'].font      = fnt(size=11)
    ws['E7'].alignment = aln('center')
    ws['E7'].border    = b_all()

    # ── FILAS 8-15: datos cliente / vendedor ────────────
    # Estructura exacta del ref: A=label, B:D merged=valor, E=label, F=valor
    # Bordes: fila 8 top+left, filas 9-14 solo left, fila 15 left+bottom
    etq_cli = [
        ("Name/Nombre",           d['cliente_nombre']),
        ("ID/Nit",                d['cliente_nit']),
        ("Address/Direccion",     d['cliente_dir']),
        ("City/Ciudad",           d['cliente_ciu']),
        ("State/Departemento",    d['cliente_dep']),
        ("Country/Pais",          d['cliente_pais']),
        ("Phone Number/Telefono", d['cliente_tel']),
        ("Contact/Contacto",      d['cliente_cont']),
    ]
    etq_ven_lbl = ["VIVELL S.AS", "ID/NIT", "Address/Direccion", "City/Ciudad",
                   "State/Departemento", "Country/Pais", "Phone Number/Telefono", "Vendor/Vendedor"]
    etq_ven_val = [d['cia_nombre'], d['cia_nit'], d['cia_dir'], d['cia_ciu'],
                   d['cia_dep'], d['cia_pais'], d['cia_tel'], d['vendedor']]

    for i, ((lbl_c, val_c), lbl_v, val_v) in enumerate(zip(etq_cli, etq_ven_lbl, etq_ven_val)):
        r = 8 + i
        is_first = (i == 0)
        is_last  = (i == 7)

        # Bordes según posición (idéntico al ref)
        if is_first:
            b_a = b_left_top(); b_b = b_ltr(); b_e = b_left_top(); b_f = b_right()
        elif is_last:
            b_a = b_lb();       b_b = b_lbr(); b_e = b_lb();       b_f = b_rb()
        else:
            b_a = b_left();     b_b = b_lr();  b_e = b_left();     b_f = b_right()

        ws[f'A{r}'].value     = lbl_c
        ws[f'A{r}'].font      = fnt(size=11)
        ws[f'A{r}'].border    = b_a

        ws.merge_cells(f'B{r}:D{r}')
        ws[f'B{r}'].value     = str(val_c)
        ws[f'B{r}'].font      = fnt(size=11)
        ws[f'B{r}'].alignment = aln('left')
        ws[f'B{r}'].border    = b_b

        ws[f'E{r}'].value     = lbl_v
        ws[f'E{r}'].font      = fnt(size=11)
        ws[f'E{r}'].border    = b_e

        ws[f'F{r}'].value     = str(val_v)
        ws[f'F{r}'].font      = fnt(size=11)
        ws[f'F{r}'].alignment = aln('left')
        ws[f'F{r}'].border    = b_f

    # ── FILA 16: Encabezado columnas tabla ──────────────
    # Fondo tema (gris oscuro), texto bold, bordes all
    FILL_HDR = PF('solid', start_color='FFD9D9D9', end_color='FFD9D9D9')
    titulos  = ["Reference/Referencia", "Size/Talla", "Colour/Color",
                "Description/Descripcion", "Composition/Composicion", "Quantity/Cantidad"]
    aligns   = [None, None, None, 'left', None, 'right']
    for col_idx, (titulo, h) in enumerate(zip(titulos, aligns), start=1):
        cell           = ws.cell(row=16, column=col_idx, value=titulo)
        cell.font      = fnt(bold=True, size=11)
        cell.fill      = FILL_HDR
        cell.border    = b_all()
        if h:
            cell.alignment = aln(h)
    ws.row_dimensions[16].height = 15

    # ── DETALLE POR CAJA ────────────────────────────────
    FILL_CAJA = PF('solid', start_color='FFD9D9D9', end_color='FFD9D9D9')
    FILL_TOT  = PF('solid', start_color='FFBFBFBF', end_color='FFBFBFBF')

    fila               = 17
    filas_cant_detalle = []
    total_gral_cant    = 0
    total_gral_pb      = 0.0
    total_gral_pn      = 0.0
    num_cajas          = 0

    for caja, items in df.groupby('Caja', sort=True):
        val_pb   = float(items['f2_ent_peso_bruto'].iloc[0]) if pd.notna(items['f2_ent_peso_bruto'].iloc[0]) else 0.0
        val_pn   = float(items['f2_ent_peso_neto'].iloc[0])  if pd.notna(items['f2_ent_peso_neto'].iloc[0])  else 0.0
        remision = str(items['f2_remision'].iloc[0])
        num_cajas += 1

        # Fila de caja (fondo gris, merged A:F, borde all)
        ws.merge_cells(f'A{fila}:F{fila}')
        c = ws[f'A{fila}']
        c.value     = f"Box/CajaN°: {caja} -  Remision/Remission: {remision}"
        c.font      = fnt(size=11)
        c.fill      = FILL_CAJA
        c.alignment = aln('center')
        c.border    = b_all()
        ws.row_dimensions[fila].height = 15
        fila += 1

        # Filas de ítems — SIN bordes, SIN relleno (igual al ref)
        cant_caja  = 0.0
        filas_caja = []
        for _, row in items.iterrows():
            v_can = float(row['f2_cantidad'])
            datos = [
                str(row['f2_referencia']),
                str(row['f2_ext2_det']),
                str(row['f2_id_ext1_det']),
                str(row['f2_desc_crit1']),
                str(row['f2_desc_crit2']),
                str(int(v_can)),
            ]
            for col_idx, valor in enumerate(datos, start=1):
                cell      = ws.cell(row=fila, column=col_idx, value=valor)
                cell.font = fnt(size=8)
                # Sin bordes, sin relleno — igual al ref
            ws.row_dimensions[fila].height = 13
            filas_caja.append(fila)
            filas_cant_detalle.append(fila)
            cant_caja += v_can
            fila      += 1

        # Subtotal caja — bold, fondo, sin bordes (igual al ref)
        pb_str = f"{val_pb:.2f}".replace('.', ',')
        pn_str = f"{val_pn:.2f}".replace('.', ',')

        ws[f'A{fila}'].value = f"Total BOX/CAJA : {caja}"
        ws[f'A{fila}'].font  = fnt(bold=True, size=9)
        ws[f'A{fila}'].fill  = FILL_TOT

        ws[f'D{fila}'].value = f"Gross Weight/Peso Bruto : {pb_str}"
        ws[f'D{fila}'].font  = fnt(bold=True, size=9)
        ws[f'D{fila}'].fill  = FILL_TOT

        ws[f'E{fila}'].value = f"Net Weight/Peso Neto    : {pn_str}"
        ws[f'E{fila}'].font  = fnt(bold=True, size=9)
        ws[f'E{fila}'].fill  = FILL_TOT

        ws[f'F{fila}'].value = f"=SUM(F{filas_caja[0]}:F{filas_caja[-1]})" if filas_caja else cant_caja
        ws[f'F{fila}'].font  = fnt(bold=True, size=9)
        ws[f'F{fila}'].fill  = FILL_TOT

        ws.row_dimensions[fila].height = 13
        fila += 1

        total_gral_cant += cant_caja
        total_gral_pb   += val_pb
        total_gral_pn   += val_pn

    # ── GRAN TOTAL ──────────────────────────────────────
    pb_tot = f"{total_gral_pb:.2f}".replace('.', ',')
    pn_tot = f"{total_gral_pn:.2f}".replace('.', ',')

    ws[f'A{fila}'].value = f"Total BOXS/CAJAS : {num_cajas}"
    ws[f'A{fila}'].font  = fnt(bold=True, size=9)
    ws[f'A{fila}'].fill  = FILL_TOT

    ws[f'D{fila}'].value = f"Gross Weight/Peso Bruto : {pb_tot}"
    ws[f'D{fila}'].font  = fnt(bold=True, size=9)
    ws[f'D{fila}'].fill  = FILL_TOT

    ws[f'E{fila}'].value = f"Net Weight/Peso Neto    : {pn_tot}"
    ws[f'E{fila}'].font  = fnt(bold=True, size=9)
    ws[f'E{fila}'].fill  = FILL_TOT

    refs = '+'.join([f'F{r}' for r in filas_cant_detalle])
    ws[f'F{fila}'].value = f"={refs}" if filas_cant_detalle else total_gral_cant
    ws[f'F{fila}'].font  = fnt(bold=True, size=9)
    ws[f'F{fila}'].fill  = FILL_TOT
    ws.row_dimensions[fila].height = 15

    # Serializar
    try:
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer.getvalue(), f"Packing_List_{consec_inicial}.xlsx"
    except Exception as e:
        print(f"Error al generar Excel: {e}")
        return None, None